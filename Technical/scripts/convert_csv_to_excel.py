#!/usr/bin/env python3
"""
CSV to Excel Converter
Converts existing CSV analysis outputs to single-sheet Excel format
Per Nick's strict requirements: ONE SHEET PER FILE, professional formatting
"""

import pandas as pd
import numpy as np
from pathlib import Path
import logging
from datetime import datetime
from typing import Dict, List, Any, Tuple
import json
import warnings
import os
DATA_ROOT = Path(os.environ.get("DATA_ROOT", "data"))

warnings.filterwarnings('ignore')

class CSVToExcelConverter:
    """
    Converts CSV files to professional Excel format
    Strict compliance with Nick's one-sheet requirement
    """

    def __init__(self, base_path: str = str(DATA_ROOT)):
        self.base_path = Path(base_path)
        self.csv_path = self.base_path / "Output" / "Data" / "analysis_outputs"
        self.excel_path = self.base_path / "Output" / "Data" / "excel_outputs"
        self.excel_path.mkdir(parents=True, exist_ok=True)

        # Configure logging
        log_file = self.excel_path / f"conversion_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)

        self.logger.info("CSV to Excel Converter Initialized")
        self.logger.info(f"CSV source: {self.csv_path}")
        self.logger.info(f"Excel output: {self.excel_path}")

        self.conversion_results = {}

    def find_csv_files(self) -> List[Path]:
        """Find all CSV files to convert"""
        csv_files = []

        # Search in analysis_outputs directory
        if self.csv_path.exists():
            csv_files.extend(list(self.csv_path.rglob("*.csv")))

        # Also search in streamlined_analysis directory
        streamlined_path = self.base_path / "Output" / "Data" / "streamlined_analysis"
        if streamlined_path.exists():
            csv_files.extend(list(streamlined_path.glob("*.csv")))

        # Remove any existing Excel files from the list
        csv_files = [f for f in csv_files if not f.name.endswith('.xlsx')]

        self.logger.info(f"Found {len(csv_files)} CSV files to convert")
        return csv_files

    def clean_column_names(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean column names to be machine-readable"""
        # Replace spaces and special characters with underscores
        df.columns = df.columns.str.replace(' ', '_')
        df.columns = df.columns.str.replace('[^a-zA-Z0-9_]', '', regex=True)
        df.columns = df.columns.str.replace('__', '_', regex=True)
        df.columns = df.columns.str.strip('_')

        # Ensure columns start with letter or underscore
        df.columns = ['col_' + col if col[0].isdigit() else col for col in df.columns]

        return df

    def format_dataframe_for_excel(self, df: pd.DataFrame) -> pd.DataFrame:
        """Format DataFrame for professional Excel output"""
        # Make a copy to avoid modifying original
        formatted_df = df.copy()

        # Clean column names
        formatted_df = self.clean_column_names(formatted_df)

        # Convert numeric columns to proper format
        for col in formatted_df.columns:
            # Try to convert to numeric
            numeric_series = pd.to_numeric(formatted_df[col], errors='coerce')
            if not numeric_series.isna().all():
                # Check if it's integer or float
                if numeric_series.dropna().apply(lambda x: x.is_integer()).all():
                    formatted_df[col] = numeric_series.astype('Int64')
                else:
                    formatted_df[col] = numeric_series.round(4)  # Round to 4 decimal places

        return formatted_df

    def validate_excel_file(self, excel_file: Path) -> Tuple[bool, str]:
        """Validate that Excel file meets Nick's requirements"""
        try:
            # Read the Excel file
            xl_file = pd.ExcelFile(excel_file)

            # Check: Exactly one sheet
            if len(xl_file.sheet_names) != 1:
                return False, f"FAILED: Multiple sheets detected ({len(xl_file.sheet_names)} sheets)"

            # Check: Machine-readable column names
            df = pd.read_excel(excel_file, sheet_name=0)
            invalid_cols = []
            for col in df.columns:
                if not isinstance(col, str) or not col.replace('_', '').replace('-', '').isalnum():
                    invalid_cols.append(col)

            if invalid_cols:
                return False, f"FAILED: Invalid column names: {invalid_cols}"

            # Check: Professional formatting (no crazy colors detected in basic check)
            # This is a simplified check - would need more sophisticated formatting validation

            return True, "PASSED: One sheet with machine-readable columns"

        except Exception as e:
            return False, f"FAILED: Error reading file - {str(e)}"

    def convert_csv_to_excel(self, csv_file: Path) -> Dict[str, Any]:
        """Convert a single CSV file to Excel format"""
        result = {
            'csv_file': str(csv_file),
            'excel_file': None,
            'status': 'failed',
            'rows': 0,
            'columns': 0,
            'validation': None,
            'error': None
        }

        try:
            self.logger.info(f"Converting: {csv_file.name}")

            # Read CSV file
            df = pd.read_csv(csv_file)
            result['rows'] = len(df)
            result['columns'] = len(df.columns)

            if df.empty:
                result['error'] = "CSV file is empty"
                return result

            # Format DataFrame
            formatted_df = self.format_dataframe_for_excel(df)

            # Create Excel filename
            excel_filename = csv_file.stem + '.xlsx'
            excel_file = self.excel_path / excel_filename

            # Write to Excel with professional formatting
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                formatted_df.to_excel(writer, sheet_name='Data', index=False)

                # Get the workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets['Data']

                # Apply professional Black & White formatting
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

                # Header formatting
                header_font = Font(bold=True, color="000000")
                header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")

                # Data formatting
                data_font = Font(color="000000")
                data_alignment = Alignment(horizontal="left", vertical="center")

                # Border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Apply formatting to headers
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Apply formatting to data
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width

            result['excel_file'] = str(excel_file)

            # Validate the created Excel file
            validation_result = self.validate_excel_file(excel_file)
            result['validation'] = validation_result

            if validation_result[0]:  # Passed validation
                result['status'] = 'success'
                self.logger.info(f"✓ Successfully converted: {excel_filename}")
            else:
                self.logger.warning(f"⚠ Conversion with issues: {excel_filename} - {validation_result[1]}")

        except Exception as e:
            result['error'] = str(e)
            self.logger.error(f"✗ Failed to convert {csv_file.name}: {str(e)}")

        return result

    def convert_all_csv_files(self) -> Dict[str, Any]:
        """Convert all CSV files to Excel format"""
        self.logger.info("Starting batch CSV to Excel conversion...")

        csv_files = self.find_csv_files()
        if not csv_files:
            self.logger.warning("No CSV files found to convert")
            return {'status': 'no_files', 'converted': 0, 'failed': 0}

        results = []
        successful_conversions = 0
        failed_conversions = 0

        for csv_file in csv_files:
            result = self.convert_csv_to_excel(csv_file)
            results.append(result)

            if result['status'] == 'success':
                successful_conversions += 1
            else:
                failed_conversions += 1

        # Store results
        self.conversion_results = {
            'conversion_timestamp': datetime.now().isoformat(),
            'total_files': len(csv_files),
            'successful_conversions': successful_conversions,
            'failed_conversions': failed_conversions,
            'success_rate': successful_conversions / len(csv_files) if csv_files else 0,
            'results': results
        }

        # Save conversion metadata
        metadata_file = self.excel_path / "conversion_metadata.json"
        with open(metadata_file, 'w') as f:
            json.dump(self.conversion_results, f, indent=2, default=str)

        self.logger.info(f"Conversion complete: {successful_conversions}/{len(csv_files)} files converted successfully")
        return self.conversion_results

    def generate_conversion_report(self) -> str:
        """Generate a conversion summary report"""
        if not self.conversion_results:
            return "No conversion results available"

        report = []
        report.append("=" * 80)
        report.append("CSV TO EXCEL CONVERSION REPORT")
        report.append("=" * 80)
        report.append(f"Conversion Date: {self.conversion_results['conversion_timestamp']}")
        report.append(f"Total Files: {self.conversion_results['total_files']}")
        report.append(f"Successful: {self.conversion_results['successful_conversions']}")
        report.append(f"Failed: {self.conversion_results['failed_conversions']}")
        report.append(f"Success Rate: {self.conversion_results['success_rate']:.1%}")
        report.append("")

        # Successful conversions
        successful = [r for r in self.conversion_results['results'] if r['status'] == 'success']
        if successful:
            report.append("SUCCESSFUL CONVERSIONS:")
            report.append("-" * 40)
            for result in successful:
                report.append(f"✓ {Path(result['csv_file']).name} → {Path(result['excel_file']).name}")
                report.append(f"  Rows: {result['rows']:,}, Columns: {result['columns']}")
                report.append(f"  Validation: {result['validation'][1]}")
                report.append("")

        # Failed conversions
        failed = [r for r in self.conversion_results['results'] if r['status'] == 'failed']
        if failed:
            report.append("FAILED CONVERSIONS:")
            report.append("-" * 40)
            for result in failed:
                report.append(f"✗ {Path(result['csv_file']).name}")
                report.append(f"  Error: {result['error']}")
                report.append("")

        report.append("=" * 80)
        report.append(f"Excel files saved to: {self.excel_path}")
        report.append("=" * 80)

        return "\n".join(report)

def main():
    """Main execution function"""
    converter = CSVToExcelConverter()
    results = converter.convert_all_csv_files()

    print("\n" + converter.generate_conversion_report())

    # Save report to file
    report_file = converter.excel_path / "conversion_report.txt"
    with open(report_file, 'w') as f:
        f.write(converter.generate_conversion_report())

    return results

if __name__ == "__main__":
    main()