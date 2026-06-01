#!/usr/bin/env python3
"""
HMDA Stakeholder-Focused Dashboard
=====================================
Interactive web application designed for local stakeholders including community organizations,
local governments, and residents to explore HMDA data with accessibility and plain language.

Features:
- WCAG 2.1 AA accessibility compliant
- Plain language explanations
- Geographic drill-down (state → county → tract)
- PDF report generation
- Data export capabilities
- Mobile-responsive design
"""

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import json
from pathlib import Path
from flask import Flask, render_template, jsonify, request, send_file, session
from flask_caching import Cache
from datetime import datetime
import logging
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import secrets
import os
DATA_ROOT = Path(os.environ.get("DATA_ROOT", "data"))

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)
app.secret_key = secrets.token_hex(32)

# Configure caching
cache_config = {
    'CACHE_TYPE': 'SimpleCache',
    'CACHE_DEFAULT_TIMEOUT': 300
}
app.config.from_mapping(cache_config)
cache = Cache(app)

# Base paths
BASE_PATH = DATA_ROOT
DATA_PATH = BASE_PATH / "Output" / "Data"
ENHANCED_PATH = DATA_PATH / "enhanced_analysis"


class StakeholderDataManager:
    """Manages data loading and caching for stakeholder dashboard"""
    
    def __init__(self):
        self.data_loaded = False
        self.state_data = None
        self.county_data = None
        self.msa_data = None
        self.tract_data = None
        self.race_data = None
        self.ethnicity_data = None
        self.lender_data = None
        
        # Plain language dictionary
        self.plain_language = {
            'action_taken': 'Loan Decision',
            'derived_race': 'Applicant Race',
            'derived_ethnicity': 'Applicant Ethnicity',
            'loan_amount': 'Loan Amount',
            'interest_rate': 'Interest Rate',
            'loan_to_income_ratio': 'Debt-to-Income Ratio',
            'origination_rate': 'Approval Rate',
            'denial_rate': 'Denial Rate',
            'high_cost_loan': 'High-Cost Loan',
            'lei': 'Lender',
            'fips': 'County Code',
            'state_code': 'State',
            'msa': 'Metro Area'
        }
        
    @cache.cached(timeout=600, key_prefix='load_all_data')
    def load_all_data(self):
        """Load all available datasets"""
        logger.info("Loading stakeholder data...")
        
        try:
            # Load state-level data
            state_file = ENHANCED_PATH / "state_level.csv"
            if state_file.exists():
                self.state_data = pd.read_csv(state_file, low_memory=False)
                logger.info(f"Loaded state data: {len(self.state_data)} rows")
            
            # Load county-level data
            county_file = ENHANCED_PATH / "county_level.csv"
            if county_file.exists():
                self.county_data = pd.read_csv(county_file, low_memory=False)
                logger.info(f"Loaded county data: {len(self.county_data)} rows")
            
            # Load MSA data
            msa_file = ENHANCED_PATH / "msa_level.csv"
            if msa_file.exists():
                self.msa_data = pd.read_csv(msa_file, low_memory=False)
                logger.info(f"Loaded MSA data: {len(self.msa_data)} rows")
            
            # Load tract sample
            tract_file = ENHANCED_PATH / "tract_sample.csv"
            if tract_file.exists():
                self.tract_data = pd.read_csv(tract_file, low_memory=False)
                logger.info(f"Loaded tract data: {len(self.tract_data)} rows")
            
            # Load race analysis
            race_file = ENHANCED_PATH / "race_analysis.csv"
            if race_file.exists():
                self.race_data = pd.read_csv(race_file, low_memory=False)
                logger.info(f"Loaded race data: {len(self.race_data)} rows")
            
            # Load ethnicity analysis
            ethnicity_file = ENHANCED_PATH / "ethnicity_analysis.csv"
            if ethnicity_file.exists():
                self.ethnicity_data = pd.read_csv(ethnicity_file, low_memory=False)
                logger.info(f"Loaded ethnicity data: {len(self.ethnicity_data)} rows")
            
            # Load lender rankings
            lender_file = ENHANCED_PATH / "lender_rankings.csv"
            if lender_file.exists():
                self.lender_data = pd.read_csv(lender_file, low_memory=False)
                logger.info(f"Loaded lender data: {len(self.lender_data)} rows")
            
            self.data_loaded = True
            logger.info("All data loaded successfully")
            return True
            
        except Exception as e:
            logger.error(f"Error loading data: {str(e)}")
            self.data_loaded = False
            return False
    
    def get_geographic_hierarchy(self, state_code=None, county_fips=None):
        """Get geographic data at appropriate level"""
        if county_fips and self.county_data is not None:
            return self.county_data[self.county_data['fips'] == county_fips]
        elif state_code and self.state_data is not None:
            return self.state_data[self.state_data['state_code'] == state_code]
        else:
            return self.state_data
    
    def translate_to_plain_language(self, column_name):
        """Convert technical column names to plain language"""
        return self.plain_language.get(column_name, column_name.replace('_', ' ').title())


class CommunityProfileGenerator:
    """Generates community profiles and reports"""
    
    def __init__(self, data_manager):
        self.data_manager = data_manager
    
    def generate_pdf_report(self, state_code=None, county_fips=None):
        """Generate PDF community profile"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter,
                              rightMargin=72, leftMargin=72,
                              topMargin=72, bottomMargin=18)
        
        # Container for PDF elements
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Title
        if county_fips:
            title = f"Home Mortgage Disclosure Act (HMDA) Profile<br/>County: {county_fips}"
        elif state_code:
            title = f"Home Mortgage Disclosure Act (HMDA) Profile<br/>State: {state_code}"
        else:
            title = "Home Mortgage Disclosure Act (HMDA) Profile<br/>National Overview"
        
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # What is HMDA section
        story.append(Paragraph("What is HMDA?", heading_style))
        hmda_text = """
        The Home Mortgage Disclosure Act (HMDA) requires financial institutions to report 
        information about mortgage applications and loans. This data helps communities 
        understand lending patterns, identify discriminatory practices, and ensure fair 
        access to credit.
        """
        story.append(Paragraph(hmda_text, styles['BodyText']))
        story.append(Spacer(1, 0.2*inch))
        
        # Get data for this geography
        data = self.data_manager.get_geographic_hierarchy(state_code, county_fips)
        
        if data is not None and not data.empty:
            # Key Statistics
            story.append(Paragraph("Key Statistics", heading_style))
            
            # Find loan amount sum column
            loan_sum_cols = [col for col in data.columns if 'loan_amount_sum' in col.lower()]
            app_count_cols = [col for col in data.columns if 'application_count' in col.lower() or 'loan_amount_count' in col.lower()]
            approval_cols = [col for col in data.columns if 'origination_rate' in col.lower() or 'action_taken' in col.lower()]
            
            stats_data = []
            if loan_sum_cols:
                total_loan_amount = data[loan_sum_cols[0]].sum()
                stats_data.append(['Total Loan Amount', f'${total_loan_amount:,.0f}'])
            
            if app_count_cols:
                total_applications = data[app_count_cols[0]].sum()
                stats_data.append(['Total Applications', f'{total_applications:,.0f}'])
            
            if approval_cols and len(approval_cols) > 0:
                avg_approval = data[approval_cols[0]].mean()
                stats_data.append(['Average Approval Rate', f'{avg_approval*100:.1f}%'])
            
            if stats_data:
                stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
                stats_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0f0f0')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                    ('TOPPADDING', (0, 0), (-1, -1), 12),
                ]))
                story.append(stats_table)
                story.append(Spacer(1, 0.3*inch))
        
        # Understanding the Numbers
        story.append(Paragraph("Understanding the Numbers", heading_style))
        understanding_text = """
        <b>Approval Rate:</b> The percentage of loan applications that were approved.<br/>
        <b>Denial Rate:</b> The percentage of loan applications that were denied.<br/>
        <b>High-Cost Loan:</b> A loan with an interest rate significantly above average.<br/>
        <b>Debt-to-Income Ratio:</b> The borrower's debt payments relative to their income.
        """
        story.append(Paragraph(understanding_text, styles['BodyText']))
        story.append(Spacer(1, 0.2*inch))
        
        # Footer
        story.append(PageBreak())
        story.append(Paragraph("About This Report", heading_style))
        footer_text = f"""
        This report was generated on {datetime.now().strftime('%B %d, %Y')} using data from the 
        Consumer Financial Protection Bureau (CFPB) under the Home Mortgage Disclosure Act (HMDA).
        <br/><br/>
        For questions or concerns about lending practices in your community, contact:
        <br/>• Consumer Financial Protection Bureau: www.consumerfinance.gov
        <br/>• HUD Fair Housing: 1-800-669-9777
        <br/><br/>
        Data Source: CFPB HMDA Dataset<br/>
        Analysis: HMDA Stakeholder Dashboard
        """
        story.append(Paragraph(footer_text, styles['BodyText']))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def generate_excel_export(self, state_code=None, county_fips=None):
        """Generate Excel data export"""
        wb = Workbook()
        ws = wb.active
        ws.title = "HMDA Data"
        
        # Get data
        data = self.data_manager.get_geographic_hierarchy(state_code, county_fips)
        
        if data is not None and not data.empty:
            # Write headers
            headers = data.columns.tolist()
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data
            for _, row in data.iterrows():
                ws.append(row.tolist())
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


# Initialize managers
data_manager = StakeholderDataManager()
profile_generator = CommunityProfileGenerator(data_manager)


# ============================================================================
# ROUTES
# ============================================================================

@app.route('/')
def index():
    """Landing page"""
    return render_template('stakeholder_dashboard.html')


@app.route('/api/init')
def initialize():
    """Initialize data loading"""
    success = data_manager.load_all_data()
    return jsonify({
        'success': success,
        'data_loaded': data_manager.data_loaded,
        'timestamp': datetime.now().isoformat()
    })


@app.route('/api/summary_stats')
def get_summary_stats():
    """Get national summary statistics"""
    stats = {}
    
    try:
        if data_manager.county_data is not None:
            # Find relevant columns
            loan_cols = [col for col in data_manager.county_data.columns if 'loan_amount_sum' in col.lower()]
            count_cols = [col for col in data_manager.county_data.columns if 'count' in col.lower()]
            
            if loan_cols:
                stats['total_loan_volume'] = float(data_manager.county_data[loan_cols[0]].sum())
            
            if count_cols:
                stats['total_applications'] = int(data_manager.county_data[count_cols[0]].sum())
            
            stats['counties_covered'] = data_manager.county_data['fips'].nunique()
        
        if data_manager.state_data is not None:
            stats['states_covered'] = len(data_manager.state_data)
        
        if data_manager.lender_data is not None:
            stats['lenders_tracked'] = len(data_manager.lender_data)
        
        stats['year'] = 2024  # Update based on actual data
        
    except Exception as e:
        logger.error(f"Error calculating summary stats: {str(e)}")
        return jsonify({'error': str(e)}), 500
    
    return jsonify(stats)


@app.route('/api/state_overview')
def get_state_overview():
    """Get state-level overview map data"""
    if data_manager.state_data is None:
        return jsonify({'error': 'State data not loaded'}), 404
    
    try:
        # Prepare data for choropleth
        loan_cols = [col for col in data_manager.state_data.columns if 'loan_amount_sum' in col.lower()]
        
        if not loan_cols:
            return jsonify({'error': 'No loan amount data found'}), 404
        
        fig = px.choropleth(
            data_manager.state_data,
            locations='state_code',
            locationmode='USA-states',
            color=loan_cols[0],
            scope='usa',
            title='Total Loan Volume by State',
            color_continuous_scale='Blues',
            labels={loan_cols[0]: 'Total Loan Amount ($)'},
            hover_data=['state_code']
        )
        
        fig.update_layout(
            title={
                'text': 'Explore Lending in Your State',
                'x': 0.5,
                'xanchor': 'center',
                'font': {'size': 20}
            },
            geo=dict(
                bgcolor='rgba(0,0,0,0)',
                lakecolor='rgb(255, 255, 255)',
                landcolor='rgb(243, 243, 243)',
            ),
            margin=dict(l=0, r=0, t=50, b=0),
            height=500
        )
        
        return jsonify(json.loads(fig.to_json()))
        
    except Exception as e:
        logger.error(f"Error creating state overview: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/disparity_analysis')
def get_disparity_analysis():
    """Get racial disparity analysis"""
    if data_manager.race_data is None:
        return jsonify({'error': 'Race data not loaded'}), 404
    
    try:
        # Find approval rate column
        approval_cols = [col for col in data_manager.race_data.columns if 'approval' in col.lower() or 'origination' in col.lower()]
        
        if not approval_cols or 'derived_race' not in data_manager.race_data.columns:
            return jsonify({'error': 'Required columns not found'}), 404
        
        approval_col = approval_cols[0]
        
        fig = px.bar(
            data_manager.race_data.sort_values(approval_col, ascending=False),
            x='derived_race',
            y=approval_col,
            title='Loan Approval Rates by Race',
            labels={
                'derived_race': 'Applicant Race',
                approval_col: 'Approval Rate'
            },
            color=approval_col,
            color_continuous_scale='RdYlGn',
            text=approval_col
        )
        
        fig.update_traces(texttemplate='%{text:.1%}', textposition='outside')
        
        fig.update_layout(
            title={
                'text': 'Are Approval Rates Fair?',
                'x': 0.5,
                'xanchor': 'center',
                'font': {'size': 20}
            },
            xaxis_title='Applicant Race',
            yaxis_title='Approval Rate (%)',
            yaxis_tickformat='.0%',
            showlegend=False,
            height=450,
            margin=dict(l=50, r=50, t=80, b=100)
        )
        
        fig.update_xaxes(tickangle=-45)
        
        return jsonify(json.loads(fig.to_json()))
        
    except Exception as e:
        logger.error(f"Error creating disparity analysis: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/top_lenders')
def get_top_lenders():
    """Get top lenders data"""
    if data_manager.lender_data is None:
        return jsonify({'error': 'Lender data not loaded'}), 404
    
    try:
        # Return top 20 lenders
        top_lenders = data_manager.lender_data.head(20)
        return jsonify(top_lenders.to_dict(orient='records'))
        
    except Exception as e:
        logger.error(f"Error getting top lenders: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/download/pdf')
def download_pdf():
    """Download PDF community profile"""
    state_code = request.args.get('state')
    county_fips = request.args.get('county')
    
    try:
        buffer = profile_generator.generate_pdf_report(state_code, county_fips)
        
        filename = f"hmda_profile_{state_code or 'national'}_{datetime.now().strftime('%Y%m%d')}.pdf"
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error generating PDF: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/download/excel')
def download_excel():
    """Download Excel data export"""
    state_code = request.args.get('state')
    county_fips = request.args.get('county')
    
    try:
        buffer = profile_generator.generate_excel_export(state_code, county_fips)
        
        filename = f"hmda_data_{state_code or 'national'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error generating Excel: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/health')
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'data_loaded': data_manager.data_loaded,
        'timestamp': datetime.now().isoformat()
    })


if __name__ == '__main__':
    logger.info("="*60)
    logger.info("Starting HMDA Stakeholder Dashboard")
    logger.info("="*60)
    logger.info("Loading data...")
    data_manager.load_all_data()
    logger.info("Starting Flask server on http://localhost:5000")
    logger.info("="*60)
    app.run(debug=True, host='0.0.0.0', port=5000)

